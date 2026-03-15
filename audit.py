"""
audit.py — Inventory Reconciliation Walking Skeleton

Pure Python, no LLM. Connects via MCP client (mcp_client.py -> mssql-mcp-server).
Runs the runbook Step 1 query to find unconsumed maintenance ticket parts, then
diagnoses each via GP qty (IV00102) and RINV history checks, classifies the error,
and writes findings to Excel with Summary + Detail tabs.

Usage:
    python audit.py
"""

import asyncio
import os
import textwrap
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

import mcp_client

load_dotenv()

# ---------------------------------------------------------------------------
# Verbose logging helpers
# ---------------------------------------------------------------------------

def log(msg: str):
    print(msg.encode("ascii", "replace").decode("ascii"), flush=True)

def log_query(label: str, database: str, sql: str):
    log(f"\n  [QUERY] {label}")
    log(f"  [DB]    {database}")
    for line in sql.splitlines():
        log(f"          {line}")

def log_result(rows: list, preview_cols: list[str] | None = None):
    log(f"  [RESULT] {len(rows)} row(s) returned")
    if rows and preview_cols:
        for i, row in enumerate(rows[:3]):
            vals = {k: row.get(k, "") for k in preview_cols if k in row}
            log(f"           row[{i}]: {vals}")
        if len(rows) > 3:
            log(f"           ... ({len(rows) - 3} more rows)")

# ---------------------------------------------------------------------------
# Queries
# ---------------------------------------------------------------------------

# Step 0: Simple connectivity check
QUERY_PING = "SELECT 1 AS ping"

# Query 1a: FAILED/STUCK tickets — drive from IntegrationTransactions (small filtered set),
# join outward to T2Online for part/ticket context.
# Covers: STUCK_PROCESSING, QTY_SHORTAGE*, QTYFULFI, OTHER categories.
# Status IDs excluded: 1=Success, 9=Cancelled.
QUERY_FAILED_TMIN = textwrap.dedent("""
SELECT
    it.CompanyDatabaseName              AS Company,
    tcp.TcaPKey                         AS TicketID,
    it.TicketLineItemID                 AS PartLineID,
    it.ItPartNumber                     AS PartNumber,
    COALESCE(tcp.TcpQuantityOrdered, it.ItQty) AS QuantityNeeded,
    it.ItOrigin                         AS Location,
    ist.IsPKey                          AS StatusID,
    ist.IsDescription                   AS StatusDescription,
    it.ItGPDocID                        AS GPDocID,
    it.ItLongError                      AS IntegrationError,
    it.it_retry_count                   AS RetryCount,
    it.ItProcessDate                    AS ProcessDate,
    it.ItPKey                           AS IntegrationID
FROM Inventory.dbo.IntegrationTransactions it
JOIN Inventory.dbo.IntegrationStatusLookup ist
    ON ist.IsPKey = it.ItIntegrationStatusID
LEFT JOIN T2Online.dbo.TicketPartsMain tcp
    ON tcp.TcpPKey = it.TicketLineItemID
WHERE it.ItGPDocID LIKE 'TMIN%'
  AND it.ItIntegrationStatusID NOT IN (1, 9)
  AND ISNULL(tcp.TcpConsumed, 0) = 0
""").strip()

# Query 1b: NOT_INTEGRATED candidates — closed-ticket parts in T2Online with
# unconsumed parts. Scoped to 30 days + closed tickets only to avoid full-table scan.
# Step 2 (batch TMIN check in Python) filters out parts that DO have TMIN records.
QUERY_NOT_INTEGRATED_CANDIDATES = textwrap.dedent("""
SELECT
    tcp.TcpPKey                     AS PartLineID,
    'UNKNOWN'                       AS Company,
    tcm.TcaPKey                     AS TicketID,
    tcp.TcpPartNumber               AS PartNumber,
    tcp.TcpQuantityOrdered          AS QuantityNeeded,
    tcp.TcpInventoryLocation        AS Location,
    NULL                            AS StatusID,
    NULL                            AS StatusDescription,
    NULL                            AS GPDocID,
    NULL                            AS IntegrationError,
    0                               AS RetryCount,
    NULL                            AS ProcessDate,
    NULL                            AS IntegrationID
FROM T2Online.dbo.TicketPartsMain tcp
JOIN T2Online.dbo.TicketCallMain tcm
    ON tcm.TcaPKey = tcp.TcaPKey
WHERE ISNULL(tcp.TcpConsumed, 0) = 0
  AND ISNULL(tcp.TcpQuantityOrdered, 0) > 0
  AND tcm.TcaStatus = 'C'
  AND tcm.TcaCallDate >= DATEADD(day, -30, GETDATE())
""").strip()

# Query 1c: Batch check — which TcpPKey values already have TMIN records?
# Called with a comma-separated list of IDs. Returns only IDs that HAVE a TMIN.
QUERY_HAS_TMIN = textwrap.dedent("""
SELECT TicketLineItemID
FROM Inventory.dbo.IntegrationTransactions
WHERE TicketLineItemID IN ({ids})
  AND ItGPDocID LIKE 'TMIN%'
GROUP BY TicketLineItemID
""").strip()

# Query 3: GP item-location qty for a specific part + location.
QUERY_GP_QTY = textwrap.dedent("""
SELECT
    ITEMNMBR,
    LOCNCODE,
    QTYONHND,
    ATYALLOC,
    QTYCOMTD
FROM IntegrationDB.dbo.IV00102
WHERE ITEMNMBR = '{part}'
  AND LOCNCODE = '{location}'
""").strip()

# Query 7: Check for RINV removal records for a specific part + location.
QUERY_RINV = textwrap.dedent("""
SELECT
    ItPKey,
    ItGPDocID,
    ItQty,
    ItIntegrationStatusID,
    ItProcessDate
FROM Inventory.dbo.IntegrationTransactions
WHERE ItGPDocID LIKE 'RINV%'
  AND ItPartNumber = '{part}'
  AND ItOrigin = '{location}'
ORDER BY ItProcessDate DESC
""").strip()

# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def classify(row: dict, gp_qty: dict, rinv_records: list) -> dict:
    """Map each ticket row to an error category + recommended action."""
    error    = row.get("IntegrationError") or ""
    status   = row.get("StatusID")
    needed   = row.get("QuantityNeeded") or 0
    retries  = row.get("RetryCount") or 0
    on_hand  = gp_qty.get("QTYONHND", 0)
    alloc    = gp_qty.get("ATYALLOC", 0)
    available = on_hand - alloc
    location = row.get("Location", "")
    high_retry = f" WARNING: High retry count ({retries}) — likely stuck for a long time." if retries >= 10 else ""

    if status is None:
        return {
            "category": "NOT_INTEGRATED",
            "action": "No integration record. Manually trigger consumption.",
        }

    if status == 5:
        return {
            "category": "STUCK_PROCESSING",
            "action": "Stuck in Processing. Check IntercompanyTransactions; may need reset.",
        }

    if "Quantity of part in ERP system is not enough" in error:
        deficit = max(0, needed - on_hand)
        if available >= needed:
            if alloc > 0:
                # Stock exists but tied up in allocation — genuine QTYFULFI lock
                return {
                    "category": "QTYFULFI",
                    "action": (
                        f"Allocation lock. QTYONHND={on_hand}, ATYALLOC={alloc}. "
                        "Check SOP10200 + Status 3 RINV records."
                    ),
                }
            else:
                # GP has enough stock, no current allocation — stale failure, safe to reprocess
                return {
                    "category": "QTYFULFI_STALE",
                    "action": (
                        f"GP has sufficient stock (QTYONHND={on_hand}, ATYALLOC=0). "
                        "Stale failure — reprocess directly."
                    ),
                }
        if rinv_records:
            return {
                "category": "QTY_SHORTAGE_RINV",
                "action": (
                    f"RINV removal likely caused shortage. Deficit={deficit}. "
                    f"Cycle Count {deficit} unit(s) at {location}, then reprocess.{high_retry}"
                ),
            }
        return {
            "category": "QTY_SHORTAGE",
            "action": (
                f"GP has {on_hand}, need {needed}, deficit={deficit}. "
                f"Investigate TINV/PINV history. Likely Cycle Count {deficit} unit(s).{high_retry}"
            ),
        }

    if "QTYFULFI" in error or "QtyShrtOpt" in error:
        if available < needed:
            deficit = max(0, needed - on_hand)
            return {
                "category": "QTY_SHORTAGE",
                "action": (
                    f"GP has {on_hand}, need {needed}, deficit={deficit}. "
                    f"QTYFULFI parameter error — investigate TINV/PINV history and Cycle Count.{high_retry}"
                ),
            }
        if alloc > 0:
            return {
                "category": "QTYFULFI",
                "action": (
                    f"Allocation lock. QTYONHND={on_hand}, ATYALLOC={alloc}. "
                    "Check SOP10200 + stuck Status 3 RINV."
                ),
            }
        return {
            "category": "QTYFULFI_STALE",
            "action": (
                f"GP has sufficient stock (QTYONHND={on_hand}, ATYALLOC=0). "
                "Stale QTYFULFI parameter error — reprocess directly."
            ),
        }

    if "Not safe to process" in error:
        return {
            "category": "NOT_SAFE",
            "action": "Integration flagged as not safe to process. Review ticket state in Trakker and GP for inconsistency before reprocessing.",
        }

    if "open" in error.lower() and "consum" in error.lower():
        return {
            "category": "TICKET_OPEN",
            "action": "Ticket is still open in Trakker — parts cannot be consumed until ticket is closed. Close or finalize the ticket, then reprocess.",
        }

    if "On Contract" in error or "move to existing location" in error:
        return {
            "category": "CONTRACT_LOCATION",
            "action": "Part is on contract and must be moved to the contract-designated location. Verify correct location then reprocess.",
        }

    return {
        "category": "OTHER",
        "action": f"Manual review. Error: {error}",
    }


_FIX_TYPE_MAP = {
    "QTYFULFI_STALE":    "RESET_TO_PENDING",
    "STUCK_PROCESSING":  "RESET_TO_PENDING",
    "QTY_SHORTAGE":      "CYCLE_COUNT_TBD",
    "QTY_SHORTAGE_RINV": "CYCLE_COUNT_TBD",
}

def get_fix_type(category: str) -> str:
    """Map error category to a triage fix-type label."""
    return _FIX_TYPE_MAP.get(category, "HUMAN_ACTION")


# ---------------------------------------------------------------------------
# MCP helpers
# ---------------------------------------------------------------------------

async def run_query(label: str, sql: str, database: str = "Inventory") -> list[dict]:
    """Execute a SELECT via MCP, log verbosely, and return list of row dicts."""
    log_query(label, database, sql)

    raw = await mcp_client.call_tool("execute_query", {"query": sql, "database": database})
    log(f"  [RAW]   {raw[:300]}{'...' if len(raw) > 300 else ''}")

    return mcp_client.parse_rows(raw)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

CATEGORY_COLORS = {
    "NOT_INTEGRATED":    "FFF2CC",  # yellow
    "STUCK_PROCESSING":  "FCE4D6",  # orange
    "QTYFULFI":          "DDEBF7",  # blue  — allocation lock, needs SOP10200 review
    "QTYFULFI_STALE":    "D9EAD3",  # green — stock available, just reprocess
    "QTY_SHORTAGE_RINV": "FFD7D7",  # red
    "QTY_SHORTAGE":      "FFD7D7",  # red
    "NOT_SAFE":          "EAD1DC",  # pink
    "TICKET_OPEN":       "FFF2CC",  # yellow
    "CONTRACT_LOCATION": "D9D9D9",  # grey
    "OTHER":             "EEEEEE",  # light grey
}

CATEGORY_FILLS = {
    cat: PatternFill("solid", fgColor=color)
    for cat, color in CATEGORY_COLORS.items()
}
_DEFAULT_FILL = PatternFill("solid", fgColor="FFFFFF")


def _header_row(ws, columns: list[str]):
    ws.append(columns)
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def write_excel(detail_rows: list[dict], filename: str):
    log(f"\n[EXCEL] Building workbook with {len(detail_rows)} detail row(s)...")
    wb = openpyxl.Workbook()

    # --- Summary tab ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _header_row(ws_sum, ["ErrorCategory", "Count"])

    counts = Counter(r["ErrorCategory"] for r in detail_rows)
    for category, count in sorted(counts.items(), key=lambda x: -x[1]):
        ws_sum.append([category, count])
        log(f"  {category}: {count}")

    # Triage summary section
    fix_counts = Counter(r["FixType"] for r in detail_rows)
    ws_sum.append([])  # blank separator row
    ws_sum.append(["Triage Summary", ""])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font = Font(bold=True)
    for fix_type in ("RESET_TO_PENDING", "CYCLE_COUNT_TBD", "HUMAN_ACTION"):
        ws_sum.append([fix_type, fix_counts.get(fix_type, 0)])
    auto_fixable = fix_counts.get("RESET_TO_PENDING", 0) + fix_counts.get("CYCLE_COUNT_TBD", 0)
    ws_sum.append(["Total Auto-Fixable", auto_fixable])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font = Font(bold=True)

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 10

    # --- Detail tab ---
    ws_det = wb.create_sheet("Detail")
    columns = [
        "Company", "TicketID", "PartLineID", "PartNumber", "QuantityNeeded", "Location",
        "ProcessDate", "DaysOpen",
        "StatusID", "StatusDescription", "ErrorCategory", "FixType",
        "GPQtyOnHand", "GPAllocated", "GPAvailable", "Deficit",
        "HasRINV", "RetryCount", "IntegrationError",
        "RecommendedAction", "IntegrationID",
    ]
    _header_row(ws_det, columns)

    for row in detail_rows:
        ws_det.append([row.get(c, "") for c in columns])
        fill = CATEGORY_FILLS.get(row.get("ErrorCategory", "OTHER"), _DEFAULT_FILL)
        for cell in ws_det[ws_det.max_row]:
            cell.fill = fill

    col_widths = {
        "Company": 14, "TicketID": 14, "PartLineID": 12, "PartNumber": 18, "QuantityNeeded": 14,
        "Location": 12, "ProcessDate": 14, "DaysOpen": 10,
        "StatusID": 10, "StatusDescription": 22, "ErrorCategory": 22, "FixType": 20,
        "GPQtyOnHand": 14, "GPAllocated": 14, "GPAvailable": 14, "Deficit": 10,
        "HasRINV": 10, "RetryCount": 12, "IntegrationError": 40,
        "RecommendedAction": 55, "IntegrationID": 16,
    }
    for i, col in enumerate(columns, 1):
        letter = openpyxl.utils.get_column_letter(i)
        ws_det.column_dimensions[letter].width = col_widths.get(col, 14)

    ws_det.freeze_panes = "A2"

    # --- Staged Fixes tab ---
    staged_rows = [r for r in detail_rows if r.get("FixType") != "HUMAN_ACTION"]
    staged_rows.sort(key=lambda r: r.get("DaysOpen") if isinstance(r.get("DaysOpen"), int) else 0, reverse=True)

    ws_fix = wb.create_sheet("Staged Fixes")
    fix_columns = [
        "Company", "TicketID", "PartLineID", "PartNumber",
        "Location", "ErrorCategory", "DaysOpen", "FixType",
    ]
    _header_row(ws_fix, fix_columns)

    for row in staged_rows:
        ws_fix.append([row.get(c, "") for c in fix_columns])
        fill = CATEGORY_FILLS.get(row.get("ErrorCategory", "OTHER"), _DEFAULT_FILL)
        for cell in ws_fix[ws_fix.max_row]:
            cell.fill = fill

    fix_col_widths = {
        "Company": 14, "TicketID": 14, "PartLineID": 12, "PartNumber": 18,
        "Location": 12, "ErrorCategory": 22, "DaysOpen": 10, "FixType": 20,
    }
    for i, col in enumerate(fix_columns, 1):
        letter = openpyxl.utils.get_column_letter(i)
        ws_fix.column_dimensions[letter].width = fix_col_widths.get(col, 14)

    ws_fix.freeze_panes = "A2"
    log(f"  Staged Fixes tab: {len(staged_rows)} auto-fixable row(s)")

    wb.save(filename)
    log(f"\n[DONE] Report written -> {filename}")


# ---------------------------------------------------------------------------
# Main audit loop
# ---------------------------------------------------------------------------

async def main():
    log("=== Inventory Reconciliation Audit ===\n")

    try:
        # ------------------------------------------------------------------
        # Step 0: Connectivity check
        # ------------------------------------------------------------------
        log("Step 0: Testing MCP server connectivity...")
        ping = await run_query("Connectivity ping", QUERY_PING, database="Inventory")
        log_result(ping)
        if not ping:
            log("\n[ERROR] MCP server returned no response to SELECT 1. Check MCP_SERVER_PATH and DB credentials.")
            return
        log("  MCP server is reachable.\n")

        # ------------------------------------------------------------------
        # Step 1a: Failed/stuck TMIN records from IntegrationTransactions.
        # ------------------------------------------------------------------
        log("Step 1a: Pulling failed/stuck TMIN records...")
        failed_tickets = await run_query("Failed/stuck TMIN records", QUERY_FAILED_TMIN, database="Inventory")
        log_result(failed_tickets, preview_cols=["Company", "TicketID", "PartNumber", "Location"])
        log(f"  Found {len(failed_tickets)} failed/stuck ticket part(s).")

        # ------------------------------------------------------------------
        # Step 1b: NOT_INTEGRATED — two-step approach to avoid cross-DB timeout.
        #   1b-i:  Get candidates from T2Online (closed tickets, 30 days).
        #   1b-ii: Batch-check which ones already have TMIN records.
        # ------------------------------------------------------------------
        log("\nStep 1b: Pulling NOT_INTEGRATED candidates (closed tickets, 30 days)...")
        candidates = await run_query(
            "NOT_INTEGRATED candidates", QUERY_NOT_INTEGRATED_CANDIDATES, database="T2Online"
        )
        log(f"  Found {len(candidates)} candidate part(s) from closed tickets.")

        not_integrated = []
        if candidates:
            # Batch check: which candidates already have a TMIN record?
            pkey_list = [str(c["PartLineID"]) for c in candidates if c.get("PartLineID")]
            # Process in batches of 500 to stay under SQL parameter limits
            has_tmin_set = set()
            for batch_start in range(0, len(pkey_list), 500):
                batch = pkey_list[batch_start:batch_start + 500]
                ids_str = ",".join(batch)
                tmin_rows = await run_query(
                    f"TMIN batch check ({batch_start+1}-{batch_start+len(batch)})",
                    QUERY_HAS_TMIN.format(ids=ids_str),
                    database="Inventory",
                )
                for r in tmin_rows:
                    has_tmin_set.add(r.get("TicketLineItemID"))

            not_integrated = [c for c in candidates if c.get("PartLineID") not in has_tmin_set]
            log(f"  {len(has_tmin_set)} candidates already have TMIN records (excluded).")

        log(f"  Found {len(not_integrated)} truly not-integrated ticket part(s).\n")

        tickets = failed_tickets + not_integrated

        if not tickets:
            log("\n[INFO] No actionable tickets found. All parts are either consumed or cancelled.")
            return

        log(f"  Total to process: {len(tickets)} ticket part(s).\n")

        # ------------------------------------------------------------------
        # Step 2 + 3: Diagnose + classify each ticket
        # ------------------------------------------------------------------
        log(f"Step 2: Running GP qty + RINV diagnostics for each ticket...")
        detail_rows = []

        for i, row in enumerate(tickets, 1):
            company   = row.get("Company", "")
            # TicketID = TicketCallMain.TcaPKey; falls back to PartLineID (TcpPKey)
            # when the IT record predates TicketLineItemID tracking or T2Online join misses.
            ticket    = row.get("TicketID") or row.get("PartLineID") or ""
            part      = row.get("PartNumber", "")
            location  = row.get("Location", "")
            needed    = row.get("QuantityNeeded") or 0
            part_line = row.get("PartLineID", "")

            log(f"\n  [{i}/{len(tickets)}] Company={company} Ticket={ticket} "
                f"Part={part} Location={location} QtyNeeded={needed}")

            # 2a+b: GP qty + RINV check in parallel (independent queries)
            part_esc = part.replace("'", "''")
            loc_esc  = location.replace("'", "''")
            gp_rows, rinv_rows = await asyncio.gather(
                run_query(
                    f"GP qty — {part} @ {location}",
                    QUERY_GP_QTY.format(part=part_esc, location=loc_esc),
                    database="IntegrationDB",
                ),
                run_query(
                    f"RINV check — {part} @ {location}",
                    QUERY_RINV.format(part=part_esc, location=loc_esc),
                    database="Inventory",
                ),
            )
            log_result(gp_rows, preview_cols=["ITEMNMBR", "LOCNCODE", "QTYONHND", "ATYALLOC"])
            log_result(rinv_rows, preview_cols=["ItGPDocID", "ItQty", "ItProcessDate"])
            gp_qty = gp_rows[0] if gp_rows else {}

            # 3: Classify
            classification = classify(row, gp_qty, rinv_rows)
            category = classification["category"]
            fix_type = get_fix_type(category)
            log(f"  [CLASSIFY] category={category}  fix_type={fix_type}")
            log(f"             action={classification['action']}")

            on_hand = gp_qty.get("QTYONHND", 0)
            alloc   = gp_qty.get("ATYALLOC", 0)
            deficit = max(0, needed - on_hand)

            # Parse ProcessDate -> DaysOpen
            raw_date = row.get("ProcessDate")
            process_date = ""
            days_open = ""
            if raw_date:
                try:
                    dt = datetime.fromisoformat(str(raw_date).replace("Z", "+00:00"))
                    process_date = dt.strftime("%Y-%m-%d")
                    naive = dt.replace(tzinfo=None) if dt.tzinfo else dt
                    days_open = (datetime.now() - naive).days
                except (ValueError, TypeError):
                    pass

            detail_rows.append({
                "Company":           company,
                "TicketID":          ticket,
                "PartLineID":        part_line,
                "PartNumber":        part,
                "QuantityNeeded":    needed,
                "Location":          location,
                "ProcessDate":       process_date,
                "DaysOpen":          days_open,
                "StatusID":          row.get("StatusID", ""),
                "StatusDescription": row.get("StatusDescription", ""),
                "ErrorCategory":     category,
                "FixType":           fix_type,
                "GPQtyOnHand":       on_hand,
                "GPAllocated":       alloc,
                "GPAvailable":       on_hand - alloc,
                "Deficit":           deficit,
                "HasRINV":           "Yes" if rinv_rows else "No",
                "RetryCount":        row.get("RetryCount", ""),
                "IntegrationError":  row.get("IntegrationError", ""),
                "RecommendedAction": classification["action"],
                "IntegrationID":     row.get("IntegrationID", ""),
            })

        log(f"\nStep 3: Classified {len(detail_rows)} row(s).\n")

        # ------------------------------------------------------------------
        # Step 4: Write Excel
        # ------------------------------------------------------------------
        log("Step 4: Writing Excel report...")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"audit_{timestamp}.xlsx")
        write_excel(detail_rows, filename)

    finally:
        log("\n[MCP] Closing server connection...")
        await mcp_client.close_session()
        log("[MCP] Connection closed.")


if __name__ == "__main__":
    asyncio.run(main())
