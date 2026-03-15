"""
investigate.py — LLM Investigation Layer (Phase 4)

Reads the most recent audit Excel's Staged Fixes tab, gathers evidence per row
via MCP (parallel SQL), checks deterministic fast-path rules, and falls back to
LLM investigation via phi4-mini for ambiguous cases. Writes investigation output
to investigation_YYYYMMDD_HHMMSS.xlsx.

Usage:
    python investigate.py
    python investigate.py path/to/audit_YYYYMMDD_HHMMSS.xlsx
"""

import asyncio
import glob
import os
import sys
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

import mcp_client
from evidence import gather_evidence, format_evidence, check_fast_path, EVIDENCE_QUERIES
from llm_utils import call_llm_single_turn, parse_verdict

load_dotenv()

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def log(msg: str):
    print(msg.encode("ascii", "replace").decode("ascii"), flush=True)


# ---------------------------------------------------------------------------
# System prompt — minimal, role + output format only (~65 tokens)
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = (
    "You are an inventory reconciliation analyst. "
    "You will be given a playbook with a decision tree, evidence from database queries, "
    "and details about a failed inventory transaction. "
    "Follow the decision tree step by step using the evidence provided. "
    "Output exactly 3 lines:\n"
    "verdict: CONFIRM|ESCALATE|RECLASSIFY\n"
    "reason: <one sentence>\n"
    "new_category: <only if RECLASSIFY>"
)


# ---------------------------------------------------------------------------
# Playbook loading
# ---------------------------------------------------------------------------

PLAYBOOK_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "playbooks")


def load_playbook(category: str) -> str | None:
    """Load a playbook text file for the given category. Returns None if not found."""
    path = os.path.join(PLAYBOOK_DIR, f"{category}.txt")
    if os.path.isfile(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return None


# ---------------------------------------------------------------------------
# Find most recent audit Excel
# ---------------------------------------------------------------------------

def find_latest_audit(explicit_path: str | None = None) -> str:
    """Find the most recent audit_*.xlsx file, or use the explicitly provided path."""
    if explicit_path and os.path.isfile(explicit_path):
        return explicit_path

    project_dir = os.path.dirname(os.path.abspath(__file__))
    pattern = os.path.join(project_dir, "audit_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(
            "No audit_*.xlsx files found. Run audit.py first."
        )
    return max(files, key=os.path.getmtime)


# ---------------------------------------------------------------------------
# Read Staged Fixes tab from audit Excel
# ---------------------------------------------------------------------------

def read_staged_fixes(path: str) -> list[dict]:
    """Read the Staged Fixes tab from an audit Excel file into a list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Staged Fixes" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"'{path}' has no 'Staged Fixes' tab. Is this a Phase 2+ audit file?")

    # Read Staged Fixes tab
    ws = wb["Staged Fixes"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    header_list = [str(h) if h else f"col_{i}" for i, h in enumerate(headers)]

    staged = []
    for row_vals in rows_iter:
        row_dict = dict(zip(header_list, row_vals))
        if not row_dict.get("PartNumber") and not row_dict.get("PartLineID"):
            continue
        staged.append(row_dict)

    # Read Detail tab from the same workbook to get full row data
    detail_map = {}
    if "Detail" in wb.sheetnames:
        ws_det = wb["Detail"]
        det_rows = ws_det.iter_rows(values_only=True)
        det_headers = next(det_rows)
        det_header_list = [str(h) if h else f"col_{i}" for i, h in enumerate(det_headers)]
        for row_vals in det_rows:
            d = dict(zip(det_header_list, row_vals))
            key = (d.get("PartLineID"), d.get("PartNumber"), d.get("Location"))
            detail_map[key] = d

    wb.close()

    # Merge detail data into staged rows
    for row in staged:
        key = (row.get("PartLineID"), row.get("PartNumber"), row.get("Location"))
        detail = detail_map.get(key, {})
        for col in ("IntegrationError", "QuantityNeeded", "GPQtyOnHand", "GPAllocated",
                     "GPAvailable", "Deficit", "HasRINV", "RetryCount", "IntegrationID",
                     "StatusID", "StatusDescription", "RecommendedAction", "ProcessDate"):
            if col not in row or row[col] is None:
                row[col] = detail.get(col, "")

    return staged


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

VERDICT_COLORS = {
    "CONFIRM":     "D9EAD3",  # green
    "ESCALATE":    "FCE4D6",  # orange
    "RECLASSIFY":  "DDEBF7",  # blue
    "UNKNOWN":     "EEEEEE",  # grey
}


def _header_row(ws, columns: list[str]):
    ws.append(columns)
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def write_investigation_excel(results: list[dict], filename: str):
    """Write investigation results to Excel with Summary + Detail tabs."""
    log(f"\n[EXCEL] Building investigation workbook with {len(results)} row(s)...")
    wb = openpyxl.Workbook()

    # --- Summary tab ---
    ws_sum = wb.active
    ws_sum.title = "Investigation Summary"

    _header_row(ws_sum, ["Verdict", "Count"])
    verdict_counts = Counter(r["LLMVerdict"] for r in results)
    for verdict in ("CONFIRM", "ESCALATE", "RECLASSIFY", "UNKNOWN"):
        count = verdict_counts.get(verdict, 0)
        ws_sum.append([verdict, count])
        fill = PatternFill("solid", fgColor=VERDICT_COLORS.get(verdict, "FFFFFF"))
        for cell in ws_sum[ws_sum.max_row]:
            cell.fill = fill

    ws_sum.append([])
    _header_row(ws_sum, ["Method", "Count"])
    method_counts = Counter(r["InvestigationMethod"] for r in results)
    for method in ("fast-path", "llm", "no-playbook"):
        ws_sum.append([method, method_counts.get(method, 0)])

    ws_sum.append([])
    _header_row(ws_sum, ["Reclassified To", "Count"])
    reclass = [r for r in results if r["LLMVerdict"] == "RECLASSIFY"]
    reclass_counts = Counter(r["LLMNewCategory"] for r in reclass)
    for cat, count in sorted(reclass_counts.items(), key=lambda x: -x[1]):
        ws_sum.append([cat, count])

    ws_sum.column_dimensions["A"].width = 24
    ws_sum.column_dimensions["B"].width = 10

    # --- Detail tab ---
    ws_det = wb.create_sheet("Investigation Detail")
    columns = [
        "Company", "TicketID", "PartLineID", "PartNumber", "Location",
        "ErrorCategory", "FixType", "DaysOpen",
        "LLMVerdict", "LLMReason", "LLMNewCategory", "InvestigationMethod",
        "QuantityNeeded", "IntegrationError", "IntegrationID",
    ]
    _header_row(ws_det, columns)

    for row in results:
        ws_det.append([row.get(c, "") for c in columns])
        verdict = row.get("LLMVerdict", "UNKNOWN")
        fill = PatternFill("solid", fgColor=VERDICT_COLORS.get(verdict, "FFFFFF"))
        for cell in ws_det[ws_det.max_row]:
            cell.fill = fill

    col_widths = {
        "Company": 14, "TicketID": 14, "PartLineID": 12, "PartNumber": 18,
        "Location": 12, "ErrorCategory": 20, "FixType": 18, "DaysOpen": 10,
        "LLMVerdict": 14, "LLMReason": 50, "LLMNewCategory": 20,
        "InvestigationMethod": 16, "QuantityNeeded": 14,
        "IntegrationError": 40, "IntegrationID": 16,
    }
    for i, col in enumerate(columns, 1):
        letter = openpyxl.utils.get_column_letter(i)
        ws_det.column_dimensions[letter].width = col_widths.get(col, 14)

    ws_det.freeze_panes = "A2"

    wb.save(filename)
    log(f"[DONE] Investigation report written -> {filename}")


# ---------------------------------------------------------------------------
# Main investigation loop
# ---------------------------------------------------------------------------

async def main():
    log("=== LLM Investigation Layer (Phase 4) ===\n")

    # Find audit file
    explicit = sys.argv[1] if len(sys.argv) > 1 else None
    try:
        audit_path = find_latest_audit(explicit)
    except FileNotFoundError as e:
        log(f"[ERROR] {e}")
        return
    log(f"[INPUT] Reading: {audit_path}")

    # Read staged fixes
    try:
        staged = read_staged_fixes(audit_path)
    except ValueError as e:
        log(f"[ERROR] {e}")
        return
    log(f"[INPUT] {len(staged)} staged fix row(s) to investigate.\n")

    if not staged:
        log("[INFO] No staged fixes to investigate.")
        return

    # Connectivity check
    log("Step 0: Testing MCP server connectivity...")
    try:
        raw = await mcp_client.call_tool("execute_query", {"query": "SELECT 1 AS ping", "database": "Inventory"})
        log("  MCP server is reachable.\n")
    except Exception as e:
        log(f"[ERROR] MCP server unreachable: {e}")
        return

    # --- Investigation loop ---
    results = []
    fast_path_count = 0
    llm_count = 0
    no_playbook_count = 0

    try:
        for i, row in enumerate(staged, 1):
            category = row.get("ErrorCategory", "OTHER")
            part = row.get("PartNumber", "?")
            location = row.get("Location", "?")
            log(f"[{i}/{len(staged)}] {category} — Part={part} Location={location}")

            # 1. Gather evidence (parallel SQL queries)
            evidence = await gather_evidence(row, category)
            evidence_labels = [f"{k}({len(v)})" for k, v in evidence.items()]
            log(f"  Evidence: {', '.join(evidence_labels)}")

            # 2. Check fast-path
            fast_result = check_fast_path(category, evidence, row)
            if fast_result:
                log(f"  FAST-PATH: {fast_result['verdict']} — {fast_result['reason']}")
                fast_path_count += 1
                results.append({
                    **row,
                    "LLMVerdict": fast_result["verdict"],
                    "LLMReason": fast_result["reason"],
                    "LLMNewCategory": fast_result.get("new_category", ""),
                    "InvestigationMethod": "fast-path",
                })
                continue

            # 3. Load playbook
            playbook = load_playbook(category)
            if not playbook:
                log(f"  NO PLAYBOOK for {category} — marking UNKNOWN")
                no_playbook_count += 1
                results.append({
                    **row,
                    "LLMVerdict": "UNKNOWN",
                    "LLMReason": f"No playbook for category {category}",
                    "LLMNewCategory": "",
                    "InvestigationMethod": "no-playbook",
                })
                continue

            # 4. Format evidence packet
            evidence_text = format_evidence(row, evidence)

            # 5. Call LLM — single turn, playbook + evidence
            user_prompt = f"{playbook}\n\n---\n\n{evidence_text}"
            log(f"  Calling LLM ({len(user_prompt)} chars)...")

            try:
                raw_output = await call_llm_single_turn(SYSTEM_PROMPT, user_prompt)
                verdict = parse_verdict(raw_output)
                log(f"  LLM: {verdict['verdict']} — {verdict['reason']}")
                llm_count += 1
            except Exception as e:
                log(f"  LLM ERROR: {e}")
                verdict = {"verdict": "UNKNOWN", "reason": f"LLM error: {e}", "new_category": ""}
                llm_count += 1

            results.append({
                **row,
                "LLMVerdict": verdict["verdict"],
                "LLMReason": verdict["reason"],
                "LLMNewCategory": verdict.get("new_category", ""),
                "InvestigationMethod": "llm",
            })

        # --- Summary ---
        log(f"\n{'='*50}")
        log(f"Investigation complete: {len(results)} row(s)")
        log(f"  Fast-path confirmed: {fast_path_count}")
        log(f"  LLM investigated:    {llm_count}")
        log(f"  No playbook:         {no_playbook_count}")

        verdict_counts = Counter(r["LLMVerdict"] for r in results)
        for v in ("CONFIRM", "ESCALATE", "RECLASSIFY", "UNKNOWN"):
            log(f"  {v}: {verdict_counts.get(v, 0)}")
        log(f"{'='*50}")

        # --- Write Excel ---
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        project_dir = os.path.dirname(os.path.abspath(__file__))
        filename = os.path.join(project_dir, f"investigation_{timestamp}.xlsx")
        write_investigation_excel(results, filename)

    finally:
        log("\n[MCP] Closing server connection...")
        await mcp_client.close_session()
        log("[MCP] Connection closed.")


if __name__ == "__main__":
    asyncio.run(main())
